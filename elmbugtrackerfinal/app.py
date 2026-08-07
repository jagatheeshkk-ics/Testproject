import csv
import io
import os
from datetime import datetime, timedelta
from zipfile import BadZipFile

from flask import Flask, abort, redirect, render_template, request, Response, url_for
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__)
database_url = os.environ.get("DATABASE_URL", "sqlite:///bugs.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-production")
db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

SEVERITIES = ["Low", "Medium", "High", "Critical"]
STATUSES = ["Open", "In Progress", "Resolved", "Closed"]


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    display_name = db.Column(db.String(100), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Bug(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False, default="")
    severity = db.Column(db.String(20), nullable=False, default="Medium")
    status = db.Column(db.String(20), nullable=False, default="Open")
    reporter_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    assignee_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    reporter = db.relationship("User", foreign_keys=[reporter_id])
    assignee = db.relationship("User", foreign_keys=[assignee_id])


with app.app_context():
    db.create_all()


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"].strip()).first()
        if user and user.check_password(request.form["password"]):
            login_user(user)
            return redirect(url_for("index"))
        return render_template("login.html", error="Invalid username or password")

    return render_template("login.html", error=None)


@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form["username"].strip()
        if User.query.filter_by(username=username).first():
            return render_template("register.html", error="That username is already taken")

        user = User(username=username, display_name=request.form.get("display_name", "").strip() or username)
        user.set_password(request.form["password"])
        db.session.add(user)
        db.session.commit()
        login_user(user)
        return redirect(url_for("index"))

    return render_template("register.html", error=None)


@app.route("/users")
@login_required
def users():
    all_users = User.query.order_by(User.display_name).all()
    bug_counts = {
        user.id: Bug.query.filter((Bug.reporter_id == user.id) | (Bug.assignee_id == user.id)).count()
        for user in all_users
    }
    return render_template("users.html", users=all_users, bug_counts=bug_counts)


@app.route("/users/new", methods=["GET", "POST"])
@login_required
def new_user():
    if request.method == "POST":
        username = request.form["username"].strip()
        if User.query.filter_by(username=username).first():
            return render_template("user_form.html", error="That username is already taken")

        user = User(username=username, display_name=request.form.get("display_name", "").strip() or username)
        user.set_password(request.form["password"])
        db.session.add(user)
        db.session.commit()
        return redirect(url_for("users"))

    return render_template("user_form.html", error=None)


@app.route("/logout")
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/account", methods=["GET", "POST"])
@login_required
def account():
    if request.method == "POST":
        if not current_user.check_password(request.form.get("current_password", "")):
            return render_template("account.html", error="Current password is incorrect", success=None)

        display_name = request.form.get("display_name", "").strip()
        if display_name:
            current_user.display_name = display_name

        new_password = request.form.get("new_password", "")
        if new_password:
            if new_password != request.form.get("confirm_password", ""):
                return render_template("account.html", error="New passwords do not match", success=None)
            current_user.set_password(new_password)

        db.session.commit()
        return render_template("account.html", error=None, success="Account details updated")

    return render_template("account.html", error=None, success=None)


def daily_bug_counts(days=14):
    today = datetime.utcnow().date()
    start = today - timedelta(days=days - 1)

    rows = (
        db.session.query(func.date(Bug.created_at), func.count(Bug.id))
        .filter(func.date(Bug.created_at) >= start.isoformat())
        .group_by(func.date(Bug.created_at))
        .all()
    )
    counts_by_day = dict(rows)

    labels, counts = [], []
    for i in range(days):
        day = start + timedelta(days=i)
        labels.append(day.strftime("%b %d"))
        counts.append(counts_by_day.get(day.isoformat(), 0))
    return labels, counts


def dashboard_stats():
    status_counts = {status: Bug.query.filter_by(status=status).count() for status in STATUSES}
    severity_counts = {severity: Bug.query.filter_by(severity=severity).count() for severity in SEVERITIES}
    daily_labels, daily_counts = daily_bug_counts()
    return {
        "total_bugs": Bug.query.count(),
        "statuses": STATUSES,
        "status_counts": status_counts,
        "severities": SEVERITIES,
        "severity_values": [severity_counts[s] for s in SEVERITIES],
        "daily_labels": daily_labels,
        "daily_counts": daily_counts,
    }


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html", **dashboard_stats())


@app.route("/")
@login_required
def index():
    status_filter = request.args.get("status", "")
    severity_filter = request.args.get("severity", "")

    query = Bug.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    if severity_filter:
        query = query.filter_by(severity=severity_filter)

    bugs = query.order_by(Bug.created_at.desc()).all()

    counts = {status: Bug.query.filter_by(status=status).count() for status in STATUSES}

    return render_template(
        "index.html",
        bugs=bugs,
        statuses=STATUSES,
        severities=SEVERITIES,
        status_filter=status_filter,
        severity_filter=severity_filter,
        counts=counts,
    )


@app.route("/bugs/new", methods=["GET", "POST"])
@login_required
def new_bug():
    users = User.query.order_by(User.display_name).all()

    if request.method == "POST":
        assignee_id = request.form.get("assignee_id") or None
        bug = Bug(
            title=request.form["title"].strip(),
            description=request.form.get("description", "").strip(),
            severity=request.form.get("severity", "Medium"),
            status=request.form.get("status", "Open"),
            reporter_id=current_user.id,
            assignee_id=assignee_id,
        )
        db.session.add(bug)
        db.session.commit()
        return redirect(url_for("index"))

    return render_template("bug_form.html", bug=None, severities=SEVERITIES, statuses=STATUSES, users=users)


IMPORT_COLUMNS = ["Title", "Description", "Severity", "Status", "Assignee"]


@app.route("/bugs/import", methods=["GET", "POST"])
@login_required
def import_bugs():
    if request.method != "POST":
        return render_template("import_bugs.html", results=None, error=None)

    upload = request.files.get("file")
    if not upload or not upload.filename:
        return render_template("import_bugs.html", results=None, error="Please choose an Excel file to upload")
    if not upload.filename.lower().endswith((".xlsx", ".xlsm")):
        return render_template("import_bugs.html", results=None, error="Please upload a .xlsx Excel file")

    try:
        workbook = load_workbook(upload.stream, data_only=True)
    except (InvalidFileException, BadZipFile):
        return render_template("import_bugs.html", results=None, error="Could not read that file as an Excel workbook")

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter, [])]
    column_index = {name: header.index(name) for name in IMPORT_COLUMNS if name in header}

    if "Title" not in column_index:
        return render_template(
            "import_bugs.html", results=None, error='The sheet needs a "Title" column header in its first row'
        )

    def cell(row, name):
        idx = column_index.get(name)
        value = row[idx] if idx is not None and idx < len(row) else None
        return str(value).strip() if value is not None else ""

    users_by_name = {}
    for user in User.query.all():
        users_by_name[user.username.lower()] = user
        users_by_name[user.display_name.lower()] = user

    created = 0
    rows = []
    for i, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None for v in row):
            continue

        notes = []
        title = cell(row, "Title")
        if not title:
            rows.append({"row": i, "title": "(blank)", "created": False, "notes": ["missing title, row skipped"]})
            continue

        severity = cell(row, "Severity").title()
        if severity not in SEVERITIES:
            notes.append(f'unknown severity "{severity}", defaulted to Medium' if severity else "no severity given, defaulted to Medium")
            severity = "Medium"

        status = cell(row, "Status").title()
        if status not in STATUSES:
            notes.append(f'unknown status "{status}", defaulted to Open' if status else "no status given, defaulted to Open")
            status = "Open"

        assignee = None
        assignee_name = cell(row, "Assignee")
        if assignee_name:
            assignee = users_by_name.get(assignee_name.lower())
            if not assignee:
                notes.append(f'unknown assignee "{assignee_name}", left unassigned')

        bug = Bug(
            title=title,
            description=cell(row, "Description"),
            severity=severity,
            status=status,
            reporter_id=current_user.id,
            assignee_id=assignee.id if assignee else None,
        )
        db.session.add(bug)
        created += 1
        rows.append({"row": i, "title": title, "created": True, "notes": notes})

    db.session.commit()
    return render_template("import_bugs.html", results={"created": created, "rows": rows}, error=None)


@app.route("/bugs/import/template")
@login_required
def import_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bugs"
    sheet.append(IMPORT_COLUMNS)
    sheet.append(["Example bug title", "What went wrong and how to reproduce it", "Medium", "Open", "Alice Chen"])
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = 28

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bug_import_template.xlsx"},
    )


@app.route("/bugs/<int:bug_id>/edit", methods=["GET", "POST"])
@login_required
def edit_bug(bug_id):
    bug = Bug.query.get_or_404(bug_id)
    users = User.query.order_by(User.display_name).all()

    if request.method == "POST":
        bug.title = request.form["title"].strip()
        bug.description = request.form.get("description", "").strip()
        bug.severity = request.form.get("severity", bug.severity)
        bug.status = request.form.get("status", bug.status)
        bug.assignee_id = request.form.get("assignee_id") or None
        db.session.commit()
        return redirect(url_for("index"))

    return render_template("bug_form.html", bug=bug, severities=SEVERITIES, statuses=STATUSES, users=users)


@app.route("/bugs/<int:bug_id>/delete", methods=["POST"])
@login_required
def delete_bug(bug_id):
    bug = Bug.query.get_or_404(bug_id)
    db.session.delete(bug)
    db.session.commit()
    return redirect(url_for("index"))


@app.route("/export/csv")
@login_required
def export_csv():
    bugs = Bug.query.order_by(Bug.created_at.desc()).all()

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        ["ID", "Title", "Description", "Severity", "Status", "Reporter", "Assignee", "Created At", "Updated At"]
    )
    for bug in bugs:
        writer.writerow(
            [
                bug.id,
                bug.title,
                bug.description,
                bug.severity,
                bug.status,
                bug.reporter.display_name,
                bug.assignee.display_name if bug.assignee else "Unassigned",
                bug.created_at.strftime("%Y-%m-%d %H:%M"),
                bug.updated_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    filename = f"bug_list_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/export/report")
@login_required
def export_report():
    bugs = Bug.query.order_by(Bug.severity.desc(), Bug.created_at.desc()).all()
    return render_template("report.html", bugs=bugs, generated_at=datetime.utcnow(), **dashboard_stats())


if __name__ == "__main__":
    app.run(debug=True)
