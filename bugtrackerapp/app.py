import csv
import io
import os
import smtplib
import uuid
from datetime import datetime, timedelta
from email.message import EmailMessage
from zipfile import BadZipFile

from flask import Flask, abort, flash, redirect, render_template, request, Response, send_from_directory, url_for
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func, inspect, text
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)
database_url = os.environ.get("DATABASE_URL", "sqlite:///bugs.db")
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
print(f"[startup] DATABASE_URL env var set: {'DATABASE_URL' in os.environ}")
print(f"[startup] Using database dialect: {database_url.split('://')[0]}")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-production")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB attachment limit
db = SQLAlchemy(app)

UPLOAD_FOLDER = os.path.join(app.instance_path, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

SEVERITIES = ["Low", "Medium", "High", "Critical"]
DEFAULT_STATUSES = ["Open", "In Progress", "Resolved", "Closed"]


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    display_name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(255), nullable=True)
    password_hash = db.Column(db.String(255), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class NotificationRecipient(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(255), unique=True, nullable=False)


class Status(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(30), unique=True, nullable=False)


class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    description = db.Column(db.Text, nullable=False, default="")


class Bug(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False, default="")
    severity = db.Column(db.String(20), nullable=False, default="Medium")
    status = db.Column(db.String(30), nullable=False, default="Open")
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=True)
    reporter_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    assignee_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    attachment_filename = db.Column(db.String(255), nullable=True)
    attachment_original_name = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    project = db.relationship("Project")
    reporter = db.relationship("User", foreign_keys=[reporter_id])
    assignee = db.relationship("User", foreign_keys=[assignee_id])


def ensure_columns(table, needed_columns):
    """Add any columns a model has gained since this database's tables were
    first created. db.create_all() only creates missing tables, never alters
    existing ones, so a table created by an older version of this app is
    missing newer columns until this runs."""
    inspector = inspect(db.engine)
    if table not in inspector.get_table_names():
        return

    existing_columns = {col["name"] for col in inspector.get_columns(table)}
    for column, column_type in needed_columns.items():
        if column not in existing_columns:
            db.session.execute(text(f'ALTER TABLE "{table}" ADD COLUMN {column} {column_type}'))
    db.session.commit()


with app.app_context():
    db.create_all()
    ensure_columns(
        "bug",
        {
            "project_id": "INTEGER",
            "attachment_filename": "VARCHAR(255)",
            "attachment_original_name": "VARCHAR(255)",
        },
    )
    ensure_columns("user", {"email": "VARCHAR(255)"})
    if Status.query.count() == 0:
        for name in DEFAULT_STATUSES:
            db.session.add(Status(name=name))
        db.session.commit()


SMTP_HOST = os.environ.get("SMTP_HOST")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USERNAME or "no-reply@example.com")
SMTP_USE_TLS = os.environ.get("SMTP_USE_TLS", "true").lower() != "false"


def send_email(to_addrs, subject, body):
    """Send a plain-text email to the given addresses. Configured entirely via
    SMTP_* environment variables; if SMTP_HOST isn't set, notifications are
    disabled and this just logs instead of failing, so the app works fine
    without email configured."""
    to_addrs = [addr for addr in to_addrs if addr]
    if not to_addrs:
        return
    if not SMTP_HOST:
        print(f"[email] SMTP not configured, skipping notification to {to_addrs}: {subject}")
        return

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = SMTP_FROM
    message["To"] = ", ".join(to_addrs)
    message.set_content(body)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as smtp:
            if SMTP_USE_TLS:
                smtp.starttls()
            if SMTP_USERNAME:
                smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
            smtp.send_message(message)
    except (smtplib.SMTPException, OSError) as exc:
        print(f"[email] Failed to send notification to {to_addrs}: {exc}")


def notify_bug_assignment(bug, base_url):
    recipients = [r.email for r in NotificationRecipient.query.all()]
    if bug.assignee and bug.assignee.email:
        recipients.append(bug.assignee.email)
    if not recipients:
        return

    subject = f"[Bug #{bug.id}] {bug.title}"
    body = (
        f"A bug has been assigned to you in ELM Bug Tracking System.\n\n"
        f"Title: {bug.title}\n"
        f"Severity: {bug.severity}\n"
        f"Status: {bug.status}\n"
        f"Project: {bug.project.name if bug.project else 'No Project'}\n"
        f"Assignee: {bug.assignee.display_name if bug.assignee else 'Unassigned'}\n"
        f"Reporter: {bug.reporter.display_name}\n\n"
        f"Description:\n{bug.description or '(none)'}\n\n"
        f"View it here: {base_url}bugs/{bug.id}/edit\n"
    )
    send_email(recipients, subject, body)


def all_statuses():
    return [s.name for s in Status.query.order_by(Status.id).all()]


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        user = User.query.filter_by(username=request.form["username"].strip()).first()
        if user and user.check_password(request.form["password"]):
            login_user(user)
            return redirect(url_for("index"))
        return render_template("login.html", error="Invalid username or password")

    return render_template("login.html", error=None)


@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form["username"].strip()
        if User.query.filter_by(username=username).first():
            return render_template("register.html", error="That username is already taken")

        user = User(
            username=username,
            display_name=request.form.get("display_name", "").strip() or username,
            email=request.form.get("email", "").strip() or None,
        )
        user.set_password(request.form["password"])
        db.session.add(user)
        db.session.commit()
        login_user(user)
        return redirect(url_for("index"))

    return render_template("register.html", error=None)


@app.route("/users")
@login_required
def users():
    all_users = User.query.order_by(User.display_name).all()
    bug_counts = {
        user.id: Bug.query.filter((Bug.reporter_id == user.id) | (Bug.assignee_id == user.id)).count()
        for user in all_users
    }
    return render_template("users.html", users=all_users, bug_counts=bug_counts)


@app.route("/users/new", methods=["GET", "POST"])
@login_required
def new_user():
    if request.method == "POST":
        username = request.form["username"].strip()
        if User.query.filter_by(username=username).first():
            return render_template("user_form.html", user=None, error="That username is already taken")

        user = User(
            username=username,
            display_name=request.form.get("display_name", "").strip() or username,
            email=request.form.get("email", "").strip() or None,
        )
        user.set_password(request.form["password"])
        db.session.add(user)
        db.session.commit()
        return redirect(url_for("users"))

    return render_template("user_form.html", user=None, error=None)


@app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == "POST":
        username = request.form["username"].strip()
        existing = User.query.filter_by(username=username).first()
        if existing and existing.id != user.id:
            return render_template("user_form.html", user=user, error="That username is already taken")

        user.username = username
        user.display_name = request.form.get("display_name", "").strip() or username
        user.email = request.form.get("email", "").strip() or None

        new_password = request.form.get("password", "").strip()
        if new_password:
            user.set_password(new_password)

        db.session.commit()
        return redirect(url_for("users"))

    return render_template("user_form.html", user=user, error=None)


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("You can't delete the account you're currently logged in as.")
        return redirect(url_for("users"))

    involved = Bug.query.filter((Bug.reporter_id == user.id) | (Bug.assignee_id == user.id)).count()
    if involved:
        flash(f'Can\'t delete "{user.display_name}" — they\'re reporter or assignee on {involved} bug(s).')
        return redirect(url_for("users"))

    db.session.delete(user)
    db.session.commit()
    return redirect(url_for("users"))


@app.route("/statuses")
@login_required
def statuses():
    all_status_rows = Status.query.order_by(Status.id).all()
    bug_counts = {status.id: Bug.query.filter_by(status=status.name).count() for status in all_status_rows}
    return render_template("statuses.html", statuses=all_status_rows, bug_counts=bug_counts)


@app.route("/statuses/new", methods=["GET", "POST"])
@login_required
def new_status():
    if request.method == "POST":
        name = request.form["name"].strip()
        if not name:
            return render_template("status_form.html", error="Status name is required")
        if Status.query.filter(Status.name.ilike(name)).first():
            return render_template("status_form.html", error="That status already exists")

        db.session.add(Status(name=name))
        db.session.commit()
        return redirect(url_for("statuses"))

    return render_template("status_form.html", error=None)


@app.route("/projects")
@login_required
def projects():
    all_projects = Project.query.order_by(Project.name).all()
    bug_counts = {project.id: Bug.query.filter_by(project_id=project.id).count() for project in all_projects}
    return render_template("projects.html", projects=all_projects, bug_counts=bug_counts)


@app.route("/projects/new", methods=["GET", "POST"])
@login_required
def new_project():
    if request.method == "POST":
        name = request.form["name"].strip()
        if not name:
            return render_template("project_form.html", error="Project name is required")
        if Project.query.filter(Project.name.ilike(name)).first():
            return render_template("project_form.html", error="That project already exists")

        db.session.add(Project(name=name, description=request.form.get("description", "").strip()))
        db.session.commit()
        return redirect(url_for("projects"))

    return render_template("project_form.html", error=None)


@app.route("/notifications")
@login_required
def notifications():
    recipients = NotificationRecipient.query.order_by(NotificationRecipient.email).all()
    return render_template("notifications.html", recipients=recipients, smtp_configured=bool(SMTP_HOST))


@app.route("/notifications/new", methods=["GET", "POST"])
@login_required
def new_notification_recipient():
    if request.method == "POST":
        email = request.form["email"].strip()
        if not email or "@" not in email:
            return render_template("notification_form.html", error="Enter a valid email address")
        if NotificationRecipient.query.filter(NotificationRecipient.email.ilike(email)).first():
            return render_template("notification_form.html", error="That email is already on the list")

        db.session.add(NotificationRecipient(email=email))
        db.session.commit()
        return redirect(url_for("notifications"))

    return render_template("notification_form.html", error=None)


@app.route("/notifications/<int:recipient_id>/delete", methods=["POST"])
@login_required
def delete_notification_recipient(recipient_id):
    recipient = NotificationRecipient.query.get_or_404(recipient_id)
    db.session.delete(recipient)
    db.session.commit()
    return redirect(url_for("notifications"))


@app.route("/logout")
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/account", methods=["GET", "POST"])
@login_required
def account():
    if request.method == "POST":
        if not current_user.check_password(request.form.get("current_password", "")):
            return render_template("account.html", error="Current password is incorrect", success=None)

        display_name = request.form.get("display_name", "").strip()
        if display_name:
            current_user.display_name = display_name

        email = request.form.get("email", "").strip()
        current_user.email = email or None

        new_password = request.form.get("new_password", "")
        if new_password:
            if new_password != request.form.get("confirm_password", ""):
                return render_template("account.html", error="New passwords do not match", success=None)
            current_user.set_password(new_password)

        db.session.commit()
        return render_template("account.html", error=None, success="Account details updated")

    return render_template("account.html", error=None, success=None)


def daily_bug_counts(days=14):
    today = datetime.utcnow().date()
    start = today - timedelta(days=days - 1)

    rows = (
        db.session.query(func.date(Bug.created_at), func.count(Bug.id))
        .filter(func.date(Bug.created_at) >= start.isoformat())
        .group_by(func.date(Bug.created_at))
        .all()
    )
    counts_by_day = dict(rows)

    labels, counts = [], []
    for i in range(days):
        day = start + timedelta(days=i)
        labels.append(day.strftime("%b %d"))
        counts.append(counts_by_day.get(day.isoformat(), 0))
    return labels, counts


def dashboard_stats():
    status_list = all_statuses()
    status_counts = {status: Bug.query.filter_by(status=status).count() for status in status_list}
    severity_counts = {severity: Bug.query.filter_by(severity=severity).count() for severity in SEVERITIES}
    daily_labels, daily_counts = daily_bug_counts()
    return {
        "total_bugs": Bug.query.count(),
        "statuses": status_list,
        "status_counts": status_counts,
        "severities": SEVERITIES,
        "severity_values": [severity_counts[s] for s in SEVERITIES],
        "daily_labels": daily_labels,
        "daily_counts": daily_counts,
    }


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html", **dashboard_stats())


@app.route("/")
@login_required
def index():
    status_filter = request.args.get("status", "")
    severity_filter = request.args.get("severity", "")
    project_filter = request.args.get("project", "")

    query = Bug.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    if severity_filter:
        query = query.filter_by(severity=severity_filter)
    if project_filter:
        query = query.filter_by(project_id=int(project_filter))

    bugs = query.order_by(Bug.created_at.desc()).all()

    status_list = all_statuses()
    counts = {status: Bug.query.filter_by(status=status).count() for status in status_list}

    return render_template(
        "index.html",
        bugs=bugs,
        statuses=status_list,
        severities=SEVERITIES,
        projects=Project.query.order_by(Project.name).all(),
        status_filter=status_filter,
        severity_filter=severity_filter,
        project_filter=project_filter,
        counts=counts,
    )


def save_attachment(upload):
    if not upload or not upload.filename:
        return None, None
    original_name = secure_filename(upload.filename)
    stored_name = f"{uuid.uuid4().hex}_{original_name}"
    upload.save(os.path.join(UPLOAD_FOLDER, stored_name))
    return stored_name, original_name


@app.route("/bugs/new", methods=["GET", "POST"])
@login_required
def new_bug():
    users = User.query.order_by(User.display_name).all()
    all_projects = Project.query.order_by(Project.name).all()

    if request.method == "POST":
        assignee_id = request.form.get("assignee_id") or None
        project_id = request.form.get("project_id") or None
        stored_name, original_name = save_attachment(request.files.get("attachment"))
        bug = Bug(
            title=request.form["title"].strip(),
            description=request.form.get("description", "").strip(),
            severity=request.form.get("severity", "Medium"),
            status=request.form.get("status", "Open"),
            project_id=project_id,
            reporter_id=current_user.id,
            assignee_id=assignee_id,
            attachment_filename=stored_name,
            attachment_original_name=original_name,
        )
        db.session.add(bug)
        db.session.commit()
        notify_bug_assignment(bug, request.url_root)
        return redirect(url_for("index"))

    return render_template(
        "bug_form.html", bug=None, severities=SEVERITIES, statuses=all_statuses(), users=users, projects=all_projects
    )


IMPORT_COLUMNS = ["Title", "Description", "Severity", "Status", "Project", "Assignee"]


@app.route("/bugs/import", methods=["GET", "POST"])
@login_required
def import_bugs():
    if request.method != "POST":
        return render_template("import_bugs.html", results=None, error=None)

    upload = request.files.get("file")
    if not upload or not upload.filename:
        return render_template("import_bugs.html", results=None, error="Please choose an Excel file to upload")
    if not upload.filename.lower().endswith((".xlsx", ".xlsm")):
        return render_template("import_bugs.html", results=None, error="Please upload a .xlsx Excel file")

    try:
        workbook = load_workbook(upload.stream, data_only=True)
    except (InvalidFileException, BadZipFile):
        return render_template("import_bugs.html", results=None, error="Could not read that file as an Excel workbook")

    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter, [])]
    column_index = {name: header.index(name) for name in IMPORT_COLUMNS if name in header}

    if "Title" not in column_index:
        return render_template(
            "import_bugs.html", results=None, error='The sheet needs a "Title" column header in its first row'
        )

    def cell(row, name):
        idx = column_index.get(name)
        value = row[idx] if idx is not None and idx < len(row) else None
        return str(value).strip() if value is not None else ""

    users_by_name = {}
    for user in User.query.all():
        users_by_name[user.username.lower()] = user
        users_by_name[user.display_name.lower()] = user

    projects_by_name = {project.name.lower(): project for project in Project.query.all()}
    valid_statuses = all_statuses()

    created = 0
    rows = []
    created_bugs = []
    for i, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None for v in row):
            continue

        notes = []
        title = cell(row, "Title")
        if not title:
            rows.append({"row": i, "title": "(blank)", "created": False, "notes": ["missing title, row skipped"]})
            continue

        severity = cell(row, "Severity").title()
        if severity not in SEVERITIES:
            notes.append(f'unknown severity "{severity}", defaulted to Medium' if severity else "no severity given, defaulted to Medium")
            severity = "Medium"

        status = cell(row, "Status").title()
        if status not in valid_statuses:
            notes.append(f'unknown status "{status}", defaulted to Open' if status else "no status given, defaulted to Open")
            status = "Open"

        project = None
        project_name = cell(row, "Project")
        if project_name:
            project = projects_by_name.get(project_name.lower())
            if not project:
                notes.append(f'unknown project "{project_name}", left unassigned')

        assignee = None
        assignee_name = cell(row, "Assignee")
        if assignee_name:
            assignee = users_by_name.get(assignee_name.lower())
            if not assignee:
                notes.append(f'unknown assignee "{assignee_name}", left unassigned')

        bug = Bug(
            title=title,
            description=cell(row, "Description"),
            severity=severity,
            status=status,
            project_id=project.id if project else None,
            reporter_id=current_user.id,
            assignee_id=assignee.id if assignee else None,
        )
        db.session.add(bug)
        created_bugs.append(bug)
        created += 1
        rows.append({"row": i, "title": title, "created": True, "notes": notes})

    db.session.commit()
    for bug in created_bugs:
        if bug.assignee_id:
            notify_bug_assignment(bug, request.url_root)
    return render_template("import_bugs.html", results={"created": created, "rows": rows}, error=None)


@app.route("/bugs/import/template")
@login_required
def import_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bugs"
    sheet.append(IMPORT_COLUMNS)
    sheet.append(
        ["Example bug title", "What went wrong and how to reproduce it", "Medium", "Open", "Website Redesign", "Alice Chen"]
    )
    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column_letter].width = 28

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bug_import_template.xlsx"},
    )


@app.route("/bugs/<int:bug_id>/edit", methods=["GET", "POST"])
@login_required
def edit_bug(bug_id):
    bug = Bug.query.get_or_404(bug_id)
    users = User.query.order_by(User.display_name).all()
    all_projects = Project.query.order_by(Project.name).all()

    if request.method == "POST":
        previous_assignee_id = bug.assignee_id

        bug.title = request.form["title"].strip()
        bug.description = request.form.get("description", "").strip()
        bug.severity = request.form.get("severity", bug.severity)
        bug.status = request.form.get("status", bug.status)
        bug.assignee_id = request.form.get("assignee_id") or None
        bug.project_id = request.form.get("project_id") or None

        stored_name, original_name = save_attachment(request.files.get("attachment"))
        if stored_name:
            if bug.attachment_filename:
                old_path = os.path.join(UPLOAD_FOLDER, bug.attachment_filename)
                if os.path.exists(old_path):
                    os.remove(old_path)
            bug.attachment_filename = stored_name
            bug.attachment_original_name = original_name

        db.session.commit()
        if bug.assignee_id and bug.assignee_id != previous_assignee_id:
            notify_bug_assignment(bug, request.url_root)
        return redirect(url_for("index"))

    return render_template(
        "bug_form.html", bug=bug, severities=SEVERITIES, statuses=all_statuses(), users=users, projects=all_projects
    )


@app.route("/bugs/<int:bug_id>/delete", methods=["POST"])
@login_required
def delete_bug(bug_id):
    bug = Bug.query.get_or_404(bug_id)
    if bug.attachment_filename:
        attachment_path = os.path.join(UPLOAD_FOLDER, bug.attachment_filename)
        if os.path.exists(attachment_path):
            os.remove(attachment_path)
    db.session.delete(bug)
    db.session.commit()
    return redirect(url_for("index"))


@app.route("/bugs/<int:bug_id>/attachment")
@login_required
def download_attachment(bug_id):
    bug = Bug.query.get_or_404(bug_id)
    if not bug.attachment_filename:
        abort(404)
    return send_from_directory(
        UPLOAD_FOLDER, bug.attachment_filename, download_name=bug.attachment_original_name, as_attachment=True
    )


@app.route("/export/csv")
@login_required
def export_csv():
    bugs = Bug.query.order_by(Bug.created_at.desc()).all()

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        ["ID", "Title", "Description", "Severity", "Status", "Project", "Reporter", "Assignee", "Created At", "Updated At"]
    )
    for bug in bugs:
        writer.writerow(
            [
                bug.id,
                bug.title,
                bug.description,
                bug.severity,
                bug.status,
                bug.project.name if bug.project else "No Project",
                bug.reporter.display_name,
                bug.assignee.display_name if bug.assignee else "Unassigned",
                bug.created_at.strftime("%Y-%m-%d %H:%M"),
                bug.updated_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    filename = f"bug_list_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        buffer.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/export/report")
@login_required
def export_report():
    bugs = Bug.query.order_by(Bug.severity.desc(), Bug.created_at.desc()).all()
    return render_template("report.html", bugs=bugs, generated_at=datetime.utcnow(), **dashboard_stats())


if __name__ == "__main__":
    app.run(debug=True)
